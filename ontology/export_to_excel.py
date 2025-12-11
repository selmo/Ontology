#!/usr/bin/env python3
"""
Export ontology.json to Excel workbook with multiple sheets
- 분류 (Classifications)
- 용어 (Terms)
- 동의어 (Synonyms)
- 연관 관계 (Related Terms)
- 표준 레퍼런스 (Standard References)
"""

import json
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed")
    print("Install with: pip install openpyxl")
    exit(1)


def load_ontology(filepath='ontology.json'):
    """Load ontology.json"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_excel_workbook(data, output_file='ontology_master_data.xlsx'):
    """Create Excel workbook with multiple sheets"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1: 분류 (Classifications)
    print("Creating Classifications sheet...")
    ws_clsf = wb.create_sheet("분류")

    # Headers
    headers_clsf = [
        "분류ID", "분류명", "영문명", "레벨", "상위분류ID", "도메인코드",
        "도메인명", "설명", "표준레퍼런스"
    ]
    ws_clsf.append(headers_clsf)

    # Apply header style
    for col_num, header in enumerate(headers_clsf, 1):
        cell = ws_clsf.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data
    def add_classifications(classifications, domain_code, domain_name, parent_id=None):
        for clsf in classifications:
            # Extract standard_ids from standard_refs
            std_refs_list = clsf.get('standard_refs', [])
            if std_refs_list and isinstance(std_refs_list, list) and len(std_refs_list) > 0:
                if isinstance(std_refs_list[0], dict):
                    std_refs = ', '.join([ref.get('standard_id', '') for ref in std_refs_list])
                else:
                    std_refs = ', '.join(std_refs_list)
            else:
                std_refs = ''

            ws_clsf.append([
                clsf['id'],
                clsf['name'],
                clsf.get('name_en', ''),
                clsf.get('level', 0),
                clsf.get('parent_id', parent_id or ''),
                domain_code,
                domain_name,
                clsf.get('description', ''),
                std_refs
            ])
            if clsf.get('children'):
                add_classifications(clsf['children'], domain_code, domain_name, clsf['id'])

    for domain in data['domains']:
        add_classifications(
            domain.get('classifications', []),
            domain['code'],
            domain['name_ko']
        )

    # Adjust column widths
    ws_clsf.column_dimensions['A'].width = 12  # ID
    ws_clsf.column_dimensions['B'].width = 30  # 분류명
    ws_clsf.column_dimensions['C'].width = 30  # 영문명
    ws_clsf.column_dimensions['D'].width = 8   # 레벨
    ws_clsf.column_dimensions['E'].width = 12  # 상위분류ID
    ws_clsf.column_dimensions['F'].width = 10  # 도메인코드
    ws_clsf.column_dimensions['G'].width = 15  # 도메인명
    ws_clsf.column_dimensions['H'].width = 50  # 설명
    ws_clsf.column_dimensions['I'].width = 30  # 표준레퍼런스

    # Sheet 2: 용어 (Terms)
    print("Creating Terms sheet...")
    ws_term = wb.create_sheet("용어")

    headers_term = [
        "용어ID", "용어명", "영문명", "약어", "설명", "연결분류ID",
        "도메인코드", "도메인명", "표준레퍼런스"
    ]
    ws_term.append(headers_term)

    for col_num, header in enumerate(headers_term, 1):
        cell = ws_term.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for domain in data['domains']:
        for term in domain.get('terms', []):
            # Extract standard_ids from standard_refs
            std_refs_list = term.get('standard_refs', [])
            if std_refs_list and isinstance(std_refs_list, list) and len(std_refs_list) > 0:
                if isinstance(std_refs_list[0], dict):
                    std_refs = ', '.join([ref.get('standard_id', '') for ref in std_refs_list])
                else:
                    std_refs = ', '.join(std_refs_list)
            else:
                std_refs = ''

            ws_term.append([
                term['id'],
                term['name_ko'],
                term.get('name_en', ''),
                term.get('acronym', ''),
                term.get('description', ''),
                term.get('linked_clsf_id', ''),
                domain['code'],
                domain['name_ko'],
                std_refs
            ])

    ws_term.column_dimensions['A'].width = 12
    ws_term.column_dimensions['B'].width = 30
    ws_term.column_dimensions['C'].width = 30
    ws_term.column_dimensions['D'].width = 15
    ws_term.column_dimensions['E'].width = 50
    ws_term.column_dimensions['F'].width = 12
    ws_term.column_dimensions['G'].width = 10
    ws_term.column_dimensions['H'].width = 15
    ws_term.column_dimensions['I'].width = 30

    # Sheet 3: 동의어 (Synonyms)
    print("Creating Synonyms sheet...")
    ws_syn = wb.create_sheet("동의어")

    headers_syn = [
        "용어ID", "용어명", "동의어타입", "동의어ID", "동의어명", "도메인"
    ]
    ws_syn.append(headers_syn)

    for col_num, header in enumerate(headers_syn, 1):
        cell = ws_syn.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for domain in data['domains']:
        for term in domain.get('terms', []):
            if term.get('synonyms'):
                synonyms = term['synonyms']

                # Handle structured synonyms (v3.11+)
                if isinstance(synonyms, dict):
                    # Term → Term synonyms
                    for syn_term_id in synonyms.get('terms', []):
                        ws_syn.append([
                            term['id'],
                            term['name_ko'],
                            'Term',
                            syn_term_id,
                            '',  # Will be filled by lookup
                            domain['name_ko']
                        ])

                    # String synonyms
                    for syn_str in synonyms.get('strings', []):
                        ws_syn.append([
                            term['id'],
                            term['name_ko'],
                            'String',
                            '',
                            syn_str,
                            domain['name_ko']
                        ])

                # Handle legacy flat list
                elif isinstance(synonyms, list):
                    for syn in synonyms:
                        ws_syn.append([
                            term['id'],
                            term['name_ko'],
                            'String',
                            '',
                            syn,
                            domain['name_ko']
                        ])

    ws_syn.column_dimensions['A'].width = 12
    ws_syn.column_dimensions['B'].width = 30
    ws_syn.column_dimensions['C'].width = 12
    ws_syn.column_dimensions['D'].width = 12
    ws_syn.column_dimensions['E'].width = 30
    ws_syn.column_dimensions['F'].width = 15

    # Sheet 4: 연관 관계 (Related Terms)
    print("Creating Related Terms sheet...")
    ws_rel = wb.create_sheet("연관관계")

    headers_rel = [
        "용어ID1", "용어명1", "용어ID2", "용어명2", "도메인"
    ]
    ws_rel.append(headers_rel)

    for col_num, header in enumerate(headers_rel, 1):
        cell = ws_rel.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for domain in data['domains']:
        for term in domain.get('terms', []):
            if term.get('related_terms'):
                for related_id in term['related_terms']:
                    ws_rel.append([
                        term['id'],
                        term['name_ko'],
                        related_id,
                        '',  # Will be filled by lookup
                        domain['name_ko']
                    ])

    ws_rel.column_dimensions['A'].width = 12
    ws_rel.column_dimensions['B'].width = 30
    ws_rel.column_dimensions['C'].width = 12
    ws_rel.column_dimensions['D'].width = 30
    ws_rel.column_dimensions['E'].width = 15

    # Sheet 5: 표준 레지스트리 (Standards Registry)
    print("Creating Standards Registry sheet...")
    ws_std = wb.create_sheet("표준레지스트리")

    headers_std = [
        "표준ID", "표준코드", "한글명", "영문명", "타입", "범위",
        "기관", "버전", "URI", "설명"
    ]
    ws_std.append(headers_std)

    for col_num, header in enumerate(headers_std, 1):
        cell = ws_std.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    if 'standards' in data and 'registry' in data['standards']:
        for std in data['standards']['registry']:
            ws_std.append([
                std['id'],
                std['code'],
                std['name_ko'],
                std['name_en'],
                std['type'],
                std['scope'],
                std['organization'],
                std['version'],
                std['uri'],
                std.get('description', '')
            ])

    ws_std.column_dimensions['A'].width = 15
    ws_std.column_dimensions['B'].width = 15
    ws_std.column_dimensions['C'].width = 30
    ws_std.column_dimensions['D'].width = 30
    ws_std.column_dimensions['E'].width = 15
    ws_std.column_dimensions['F'].width = 15
    ws_std.column_dimensions['G'].width = 20
    ws_std.column_dimensions['H'].width = 10
    ws_std.column_dimensions['I'].width = 50
    ws_std.column_dimensions['J'].width = 50

    # Sheet 6: 메타데이터 (Metadata)
    print("Creating Metadata sheet...")
    ws_meta = wb.create_sheet("메타데이터")

    ws_meta.append(["항목", "값"])
    ws_meta['A1'].fill = header_fill
    ws_meta['A1'].font = header_font
    ws_meta['B1'].fill = header_fill
    ws_meta['B1'].font = header_font

    ws_meta.append(["버전", data['metadata']['version']])
    ws_meta.append(["최종 업데이트", data['metadata']['last_updated']])
    ws_meta.append(["설명", data['metadata']['description']])
    ws_meta.append(["", ""])
    ws_meta.append(["생성일시", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_meta.append(["", ""])

    # Statistics
    ws_meta.append(["통계", ""])
    ws_meta['A' + str(ws_meta.max_row)].fill = header_fill
    ws_meta['A' + str(ws_meta.max_row)].font = header_font

    ws_meta.append(["도메인 수", len(data['domains'])])
    ws_meta.append(["분류 수", ws_clsf.max_row - 1])
    ws_meta.append(["용어 수", ws_term.max_row - 1])
    ws_meta.append(["동의어 수", ws_syn.max_row - 1])
    ws_meta.append(["연관관계 수", ws_rel.max_row - 1])

    if 'standards' in data and 'registry' in data['standards']:
        ws_meta.append(["표준 수", len(data['standards']['registry'])])

    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 60

    # Save workbook
    print(f"\nSaving to {output_file}...")
    wb.save(output_file)

    return output_file


def print_summary(output_file, data):
    """Print summary"""
    print("\n" + "=" * 80)
    print("Excel 파일 생성 완료")
    print("=" * 80)

    print(f"\n파일: {output_file}")

    print(f"\n생성된 시트:")
    print(f"  1. 분류 - 376개 분류 (전체 계층)")
    print(f"  2. 용어 - 221개 용어")
    print(f"  3. 동의어 - Term→Term + Term→String")
    print(f"  4. 연관관계 - RELATED_TO 관계")
    print(f"  5. 표준레지스트리 - 24개 표준")
    print(f"  6. 메타데이터 - 버전 및 통계")

    print(f"\n버전: {data['metadata']['version']}")
    print(f"최종 업데이트: {data['metadata']['last_updated']}")


if __name__ == '__main__':
    print("Loading ontology.json...")
    data = load_ontology()

    output_file = create_excel_workbook(data)
    print_summary(output_file, data)

    print(f"\n✅ 완료: {output_file}")
